"""★ 엑셀 출력 규격 검증 (내부망 관리도구와 반드시 일치).

최소한 아래 4가지를 강제한다.
  (a) 파일명 접두사  CVE_YYYYMMDD.xls
  (b) 시트명         CVE_LIST / META
  (c) 헤더 16개      순서·철자
  (d) 전 셀 문자열   숫자·날짜형으로 새면 내부망 POI 파싱에서 깨진다

(d) 는 두 겹으로 검증한다.
  1. build_rows() 결과의 모든 값이 str 인지  (생성 단계)
  2. 저장된 .xlsx 를 openpyxl 로 다시 열어 전 셀의 타입이 문자열인지 (기록 결과)
  3. xlrd 가 설치돼 있으면 .xls 정본도 같은 방식으로 재검증 (선택)
"""
from __future__ import annotations

import csv
import re
from datetime import datetime

import pytest
from openpyxl import load_workbook

from src.aicve.excel import (
    HEADERS,
    META_SHEET_NAME,
    SHEET_NAME,
    build_file_name,
    build_meta,
    build_rows,
    cell_text,
    cleanup_old_files,
    finding_to_row,
    score_text,
    select_rows,
    sheet_names_for,
    write_excel,
)
from src.aicve.normalize import Finding
from src.aicve.scope import resolve_scope

from .test_scope import SETTINGS, WATCHLIST

RUN_ID = "20260814090000"
RUN_DATE = "20260814"


def scope_for(cli=None):
    return resolve_scope(cli=cli or {"preset": "daily"},
                         settings=SETTINGS, watchlist=WATCHLIST)


def sample_findings():
    """규격상 까다로운 값들을 일부러 섞는다 (숫자로 보이는 버전, None, 개행 등)."""
    return [
        Finding(cve_id="CVE-2026-0001", sw_name="PyTorch", source="NVD+OSV",
                vendor="PyTorch Foundation", ecosystem="pypi",
                affected_range=">=1.0.0,<1.4.2", fixed_version="1.10",
                severity="CRITICAL", cvss_score=9.8,
                cvss_vector="CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:H",
                published_date="20260801", modified_date="20260810", kev_yn="Y",
                summary="원격 코드 실행\n취약점\t설명", reference_url="https://nvd.nist.gov/x",
                collected_at="20260814090000"),
        Finding(cve_id="CVE-2026-0002", sw_name="vLLM", source="OSV",
                vendor="vLLM Project", ecosystem="pypi",
                affected_range="<0.6.0", fixed_version="0.6.0",
                severity="HIGH", cvss_score=7.5, cvss_vector="CVSS:3.1/AV:N",
                published_date="20260805", modified_date="20260811", kev_yn="N",
                summary="서비스 거부", reference_url="https://osv.dev/x",
                collected_at="20260814090000"),
        # 값이 비거나 None 인 최악의 경우
        Finding(cve_id="CVE-2026-0003", sw_name="Gradio", source="GHSA",
                vendor="", ecosystem="pypi", affected_range="*", fixed_version="",
                severity="NONE", cvss_score=None, cvss_vector="",
                published_date="", modified_date="", kev_yn="N",
                summary="[버전범위 불명확] 정보 없음", reference_url="",
                collected_at="20260814090000"),
    ]


@pytest.fixture
def built(tmp_path):
    findings = sample_findings()
    return write_excel(findings, scope_for(), RUN_ID, out_dir=tmp_path,
                       new_cnt=2, updated_cnt=1, truncated_cnt=7, run_date=RUN_DATE)


# ==========================================================================
#  (a) 파일명
# ==========================================================================
def test_file_name_prefix_and_pattern():
    name = build_file_name(RUN_DATE)
    assert name == "CVE_20260814.xls"
    assert name.startswith("CVE_"), "내부망 도구가 파일명 접두사를 검증한다"
    assert re.fullmatch(r"CVE_\d{8}\.xls", name)


def test_file_name_defaults_to_today():
    assert build_file_name() == f"CVE_{datetime.now():%Y%m%d}.xls"


def test_generated_files_named_correctly(built):
    assert built.xls_path.name == "CVE_20260814.xls"
    assert built.xlsx_path.name == "CVE_20260814.xlsx"
    assert built.csv_path.name == "CVE_20260814.csv"
    for path in built.paths():
        assert path.exists() and path.stat().st_size > 0
        assert path.name.startswith("CVE_")


# ==========================================================================
#  (b) 시트명
# ==========================================================================
def test_sheet_names(built):
    book = load_workbook(built.xlsx_path)
    assert book.sheetnames[0] == "CVE_LIST", "시트1 이름은 CVE_LIST 고정"
    assert "META" in book.sheetnames
    assert SHEET_NAME == "CVE_LIST" and META_SHEET_NAME == "META"


def test_sheet_split_names():
    assert sheet_names_for(1) == ["CVE_LIST"]
    assert sheet_names_for(3) == ["CVE_LIST", "CVE_LIST_2", "CVE_LIST_3"]


# ==========================================================================
#  (c) 헤더 16개 — 순서·철자
# ==========================================================================
EXPECTED_HEADERS = [
    "CVE_ID", "SW_NAME", "VENDOR", "AFFECTED_RANGE", "FIXED_VERSION", "SEVERITY",
    "CVSS_SCORE", "CVSS_VECTOR", "PUBLISHED_DATE", "MODIFIED_DATE", "KEV_YN",
    "SUMMARY", "REFERENCE_URL", "SOURCE", "ECOSYSTEM", "COLLECTED_AT",
]


def test_header_constant_matches_spec():
    assert list(HEADERS) == EXPECTED_HEADERS
    assert len(HEADERS) == 16


def test_header_row_written(built):
    sheet = load_workbook(built.xlsx_path)[SHEET_NAME]
    written = [sheet.cell(row=1, column=i + 1).value for i in range(16)]
    assert written == EXPECTED_HEADERS
    assert sheet.cell(row=1, column=17).value is None, "17번째 컬럼이 있으면 안 된다"


def test_csv_header_matches(built):
    with built.csv_path.open(encoding="utf-8-sig", newline="") as fp:
        assert next(csv.reader(fp)) == EXPECTED_HEADERS


def test_row_column_order_matches_headers():
    finding = sample_findings()[0]
    row = dict(zip(HEADERS, finding_to_row(finding)))
    assert row["CVE_ID"] == "CVE-2026-0001"
    assert row["SW_NAME"] == "PyTorch"
    assert row["AFFECTED_RANGE"] == ">=1.0.0,<1.4.2"
    assert row["FIXED_VERSION"] == "1.10"      # 1.1 로 깨지면 안 된다
    assert row["SEVERITY"] == "CRITICAL"
    assert row["CVSS_SCORE"] == "9.8"
    assert row["KEV_YN"] == "Y"
    assert row["SOURCE"] == "NVD+OSV"
    assert row["ECOSYSTEM"] == "pypi"
    assert row["COLLECTED_AT"] == "20260814090000"


# ==========================================================================
#  (d) 전 셀 문자열
# ==========================================================================
def test_build_rows_all_strings():
    rows = build_rows(sample_findings())
    for row in rows:
        assert len(row) == 16
        for value in row:
            assert isinstance(value, str), f"문자열이 아닌 값: {value!r} ({type(value)})"


def test_xlsx_all_cells_are_text(built):
    """저장된 파일을 다시 열어 전 셀이 문자열인지 확인한다."""
    book = load_workbook(built.xlsx_path)
    for name in book.sheetnames:
        sheet = book[name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                assert isinstance(cell.value, str), (
                    f"{name}!{cell.coordinate} 가 문자열이 아님: "
                    f"{cell.value!r} ({type(cell.value)})")
                assert cell.number_format == "@", f"{name}!{cell.coordinate} 서식이 텍스트가 아님"


def test_xls_all_cells_are_text(built):
    """정본 .xls 도 같은 방식으로 재검증 (xlrd 가 있을 때만)."""
    xlrd = pytest.importorskip("xlrd", reason="xlrd 미설치 — .xls 재검증 생략")
    book = xlrd.open_workbook(str(built.xls_path))
    assert book.sheet_names()[0] == "CVE_LIST"
    assert "META" in book.sheet_names()
    sheet = book.sheet_by_name("CVE_LIST")
    assert [sheet.cell_value(0, c) for c in range(16)] == EXPECTED_HEADERS
    for name in book.sheet_names():
        current = book.sheet_by_name(name)
        for r in range(current.nrows):
            for c in range(current.ncols):
                cell = current.cell(r, c)
                assert cell.ctype in (xlrd.XL_CELL_TEXT, xlrd.XL_CELL_EMPTY), (
                    f"{name}!{r},{c} 이 문자열 셀이 아님: ctype={cell.ctype} "
                    f"value={cell.value!r}")


def test_none_and_nan_become_empty_string():
    assert cell_text(None) == ""
    assert cell_text(float("nan")) == ""
    assert cell_text("None") == ""
    assert cell_text("") == ""
    assert score_text(None) == ""
    assert score_text("") == ""
    assert score_text("이상한값") == ""


def test_score_always_one_decimal():
    assert score_text(9.8) == "9.8"
    assert score_text(10) == "10.0"
    assert score_text(0) == "0.0"
    assert score_text("7") == "7.0"


def test_newlines_and_tabs_replaced():
    row = finding_to_row(sample_findings()[0])
    summary = row[HEADERS.index("SUMMARY")]
    assert "\n" not in summary and "\t" not in summary and "\r" not in summary
    assert summary == "원격 코드 실행 취약점 설명"


EMPTY_COLUMNS = ("CVSS_SCORE", "FIXED_VERSION", "VENDOR", "PUBLISHED_DATE",
                 "MODIFIED_DATE", "CVSS_VECTOR", "REFERENCE_URL")


def test_empty_values_are_blank_never_zero_or_none_literal(built):
    """빈 값은 '빈 셀' 이어야 한다. 0 이나 'None' 문자열이 새어 나오면 안 된다.

    (엑셀 파일에서 빈 문자열은 빈 셀로 저장된다 — openpyxl 은 None,
     xlrd 는 XL_CELL_EMPTY 로 읽는다. 둘 다 정상이며 POI 는 "" 로 받는다.)
    """
    sheet = load_workbook(built.xlsx_path)[SHEET_NAME]
    row = [c.value for c in sheet[4]]          # CVE-2026-0003 (빈 값이 많은 행)
    for column in EMPTY_COLUMNS:
        value = row[HEADERS.index(column)]
        assert value in (None, ""), f"{column} 이 비어 있지 않다: {value!r}"

    # CSV 에서는 실제로 빈 문자열로 기록된다
    with built.csv_path.open(encoding="utf-8-sig", newline="") as fp:
        third = list(csv.DictReader(fp))[2]
    assert third["CVE_ID"] == "CVE-2026-0003"
    for column in EMPTY_COLUMNS:
        assert third[column] == "", f"CSV {column} 이 비어 있지 않다: {third[column]!r}"


# ==========================================================================
#  정렬 / 절삭 / META
# ==========================================================================
def test_sort_order_severity_then_score(built):
    sheet = load_workbook(built.xlsx_path)[SHEET_NAME]
    ids = [sheet.cell(row=r, column=1).value for r in range(2, 5)]
    assert ids == ["CVE-2026-0001", "CVE-2026-0002", "CVE-2026-0003"]


def test_sort_is_stable_across_equal_severity():
    findings = [
        Finding(cve_id="CVE-2026-0009", sw_name="zSW", source="OSV",
                severity="HIGH", cvss_score=7.5),
        Finding(cve_id="CVE-2026-0008", sw_name="aSW", source="OSV",
                severity="HIGH", cvss_score=7.5),
        Finding(cve_id="CVE-2026-0007", sw_name="aSW", source="OSV",
                severity="HIGH", cvss_score=9.1),
    ]
    order = [r[0] for r in build_rows(findings)]
    assert order == ["CVE-2026-0007", "CVE-2026-0008", "CVE-2026-0009"]


def test_select_rows_truncates_by_priority():
    findings = [
        Finding(cve_id="CVE-2026-0001", sw_name="A", source="OSV",
                severity="LOW", cvss_score=2.0),
        Finding(cve_id="CVE-2026-0002", sw_name="B", source="OSV",
                severity="CRITICAL", cvss_score=9.9),
        Finding(cve_id="CVE-2026-0003", sw_name="C", source="OSV",
                severity="HIGH", cvss_score=8.0),
    ]
    kept, truncated = select_rows(findings, max_rows=2)
    assert [f.cve_id for f in kept] == ["CVE-2026-0002", "CVE-2026-0003"]
    assert truncated == 1


def test_select_rows_no_truncation():
    findings = sample_findings()
    kept, truncated = select_rows(findings, max_rows=300)
    assert len(kept) == 3 and truncated == 0


META_REQUIRED = ["RUN_ID", "COLLECTED_AT", "SCOPE_DESC", "DATE_FROM", "DATE_TO",
                 "MIN_SEVERITY", "GROUPS", "SW_NAMES", "ONLY_IN_USE", "SOURCES",
                 "EXCEL_SCOPE", "MAX_ROWS", "TRUNCATED_CNT", "TOTAL_CNT",
                 "NEW_CNT", "UPDATED_CNT", "TOOL_VERSION"]


def test_meta_keys_and_order():
    meta = build_meta(scope_for(), RUN_ID, total_cnt=3, new_cnt=2,
                      updated_cnt=1, truncated_cnt=7)
    assert [k for k, _ in meta] == META_REQUIRED
    assert all(isinstance(k, str) and isinstance(v, str) for k, v in meta)


def test_meta_sheet_written(built):
    sheet = load_workbook(built.xlsx_path)[META_SHEET_NAME]
    assert [sheet.cell(row=1, column=c).value for c in (1, 2)] == ["KEY", "VALUE"]
    values = {sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
              for r in range(2, sheet.max_row + 1)}
    assert values["RUN_ID"] == RUN_ID
    assert values["TRUNCATED_CNT"] == "7"
    assert values["TOTAL_CNT"] == "3"
    assert values["NEW_CNT"] == "2"
    assert values["UPDATED_CNT"] == "1"
    assert values["SCOPE_DESC"].startswith("SCOPE ")     # 내부망 담당자용 수집조건
    assert len(values["DATE_FROM"]) == 8


def test_meta_scope_desc_matches_scope_object(built):
    sheet = load_workbook(built.xlsx_path)[META_SHEET_NAME]
    values = {sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
              for r in range(2, sheet.max_row + 1)}
    assert values["SCOPE_DESC"] == scope_for().desc


def test_result_counts(built):
    assert built.row_count == 3
    assert built.truncated_cnt == 7
    assert built.sheet_names == ["CVE_LIST"]


# ==========================================================================
#  보관기간 정리
# ==========================================================================
def test_cleanup_old_files(tmp_path):
    for name in ("CVE_20260101.xls", "CVE_20260813.xls", "CVE_20260814.xlsx",
                 "mail_preview_20260101090000.html", "mail_preview_20260813090000.txt",
                 "keep_me.txt", "CVE_bad.xls"):
        (tmp_path / name).write_text("x", encoding="utf-8")
    removed = cleanup_old_files(tmp_path, retention_days=90,
                                today=datetime(2026, 8, 14))
    assert sorted(removed) == ["CVE_20260101.xls", "mail_preview_20260101090000.html"]
    assert (tmp_path / "CVE_20260813.xls").exists()
    assert (tmp_path / "mail_preview_20260813090000.txt").exists()
    assert (tmp_path / "keep_me.txt").exists()
    assert (tmp_path / "CVE_bad.xls").exists()


def test_cleanup_disabled_when_retention_zero(tmp_path):
    (tmp_path / "CVE_20200101.xls").write_text("x", encoding="utf-8")
    assert cleanup_old_files(tmp_path, retention_days=0) == []
    assert (tmp_path / "CVE_20200101.xls").exists()
