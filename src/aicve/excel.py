"""엑셀 출력 — 내부망 관리도구 반입용 정본.

★ 이 파일의 규격은 내부망 관리도구(Apache POI 파서)와 1바이트도 어긋나면 안 된다.
   헤더 이름·순서·시트명·파일명 접두사를 바꾸면 반입이 실패한다.

  - 파일명 : CVE_YYYYMMDD.xls  (대문자 CVE 로 시작)
  - 정본   : BIFF8 .xls (xlwt).  부가로 .xlsx(openpyxl) / .csv(UTF-8 BOM) 도 만든다.
  - 시트1  : CVE_LIST — 1행 헤더(16개 고정), 2행부터 데이터
  - 시트2  : META     — KEY | VALUE
  - ★ 모든 셀을 문자열로 기록한다.
      숫자형으로 쓰면 버전 1.10 → 1.1, 날짜 → 시리얼값으로 깨져 POI 파싱에서 오류가 난다.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import xlwt
from openpyxl import Workbook as XlsxWorkbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .logutil import get_logger
from .normalize import Finding, sort_findings

log = get_logger("excel")

TOOL_VERSION = "1.0.0"

# ==========================================================================
#  ★ 고정 규격 — 변경 금지
# ==========================================================================
SHEET_NAME = "CVE_LIST"
META_SHEET_NAME = "META"
FILE_PREFIX = "CVE_"

HEADERS: Tuple[str, ...] = (
    "CVE_ID",           # 1
    "SW_NAME",          # 2
    "VENDOR",           # 3
    "AFFECTED_RANGE",   # 4
    "FIXED_VERSION",    # 5
    "SEVERITY",         # 6
    "CVSS_SCORE",       # 7
    "CVSS_VECTOR",      # 8
    "PUBLISHED_DATE",   # 9
    "MODIFIED_DATE",    # 10
    "KEV_YN",           # 11
    "SUMMARY",          # 12
    "REFERENCE_URL",    # 13
    "SOURCE",           # 14
    "ECOSYSTEM",        # 15
    "COLLECTED_AT",     # 16
)

# 컬럼 폭 (글자 수 기준)
COLUMN_WIDTHS = (16, 22, 20, 34, 14, 10, 11, 42, 15, 15, 8, 80, 46, 12, 11, 16)

XLWT_MAX_ROWS = 65536          # xlwt(BIFF8) 시트당 행 한계 (헤더 포함)
MAX_DATA_ROWS_PER_SHEET = XLWT_MAX_ROWS - 1


@dataclass
class ExcelResult:
    xls_path: Path
    xlsx_path: Optional[Path] = None
    csv_path: Optional[Path] = None
    row_count: int = 0
    truncated_cnt: int = 0
    sheet_names: List[str] = field(default_factory=list)

    @property
    def file_name(self) -> str:
        return self.xls_path.name

    def paths(self) -> List[Path]:
        return [p for p in (self.xls_path, self.xlsx_path, self.csv_path) if p]


# ==========================================================================
#  값 변환 — 전부 문자열
# ==========================================================================
def cell_text(value: Any) -> str:
    """어떤 값이든 엑셀에 넣을 문자열로. None/NaN 은 빈 문자열."""
    if value is None:
        return ""
    if isinstance(value, float):
        # NaN 판정 (float('nan') != 자기 자신)
        if value != value:
            return ""
        return f"{value:.1f}"
    text = str(value)
    if text.strip().lower() in ("none", "nan"):
        return ""
    # 개행·탭은 공백으로 (POI 파싱 시 행이 깨지는 것을 막는다)
    return text.replace("\r\n", " ").replace("\n", " ").replace("\r", " ").replace("\t", " ")


def score_text(score: Any) -> str:
    """CVSS_SCORE 는 '0.0'~'10.0' 형태의 문자열."""
    if score is None or score == "":
        return ""
    try:
        return f"{float(score):.1f}"
    except (TypeError, ValueError):
        return ""


def finding_to_row(finding: Finding) -> List[str]:
    """Finding → CVE_LIST 한 행 (헤더 순서와 정확히 일치)."""
    return [
        cell_text(finding.cve_id),
        cell_text(finding.sw_name),
        cell_text(finding.vendor),
        cell_text(finding.affected_range or "*"),
        cell_text(finding.fixed_version),
        cell_text(finding.severity or "NONE"),
        score_text(finding.cvss_score),
        cell_text(finding.cvss_vector),
        cell_text(finding.published_date),
        cell_text(finding.modified_date),
        cell_text(finding.kev_yn or "N"),
        cell_text(finding.summary),
        cell_text(finding.reference_url),
        cell_text(finding.source),
        cell_text(finding.ecosystem or "other"),
        cell_text(finding.collected_at),
    ]


def build_rows(findings: Sequence[Finding]) -> List[List[str]]:
    """정렬 후 전 행을 문자열 리스트로. (테스트가 이 함수로 전 셀 타입을 검증한다)"""
    return [finding_to_row(f) for f in sort_findings(findings)]


def select_rows(findings: Sequence[Finding],
                max_rows: int) -> Tuple[List[Finding], int]:
    """정렬 후 max_rows 까지만 남긴다. 잘려나간 건수를 함께 돌려준다.

    정렬: SEVERITY(CRITICAL→NONE) → CVSS_SCORE 내림차순 → SW_NAME → CVE_ID
    """
    ordered = sort_findings(findings)
    if max_rows and len(ordered) > max_rows:
        return ordered[:max_rows], len(ordered) - max_rows
    return ordered, 0


def build_file_name(run_date: Optional[str] = None, suffix: str = ".xls") -> str:
    """CVE_YYYYMMDD.xls — 반드시 대문자 CVE 로 시작한다."""
    stamp = run_date or datetime.now().strftime("%Y%m%d")
    return f"{FILE_PREFIX}{stamp}{suffix}"


def build_meta(scope, run_id: str, total_cnt: int, new_cnt: int,
               updated_cnt: int, truncated_cnt: int,
               collected_at: Optional[str] = None,
               extra: Optional[List[Tuple[str, str]]] = None) -> List[Tuple[str, str]]:
    """META 시트 항목 (KEY, VALUE) — 순서 고정."""
    meta: List[Tuple[str, str]] = [
        ("RUN_ID", run_id),
        ("COLLECTED_AT", collected_at or datetime.now().strftime("%Y%m%d%H%M%S")),
    ]
    meta.extend(scope.meta_items())          # SCOPE_DESC ~ MAX_ROWS
    meta.extend([
        ("TRUNCATED_CNT", str(truncated_cnt)),
        ("TOTAL_CNT", str(total_cnt)),
        ("NEW_CNT", str(new_cnt)),
        ("UPDATED_CNT", str(updated_cnt)),
        ("TOOL_VERSION", TOOL_VERSION),
    ])
    if extra:
        meta.extend(extra)
    return [(cell_text(k), cell_text(v)) for k, v in meta]


def _chunk(rows: List[List[str]], size: int) -> List[List[List[str]]]:
    if len(rows) <= size:
        return [rows]
    return [rows[i:i + size] for i in range(0, len(rows), size)]


def sheet_names_for(chunks: int) -> List[str]:
    """CVE_LIST, CVE_LIST_2, CVE_LIST_3 …"""
    return [SHEET_NAME if i == 0 else f"{SHEET_NAME}_{i + 1}" for i in range(chunks)]


# ==========================================================================
#  .xls (BIFF8, xlwt) — 정본
# ==========================================================================
def _xlwt_styles() -> Tuple[xlwt.XFStyle, xlwt.XFStyle]:
    header = xlwt.easyxf(
        "font: bold on, colour white;"
        "pattern: pattern solid, fore_colour dark_blue;"
        "align: vert centre, horiz centre;"
        "borders: left thin, right thin, top thin, bottom thin;")
    body = xlwt.easyxf("align: vert top;")
    # 값을 반드시 텍스트로 기록하도록 서식도 텍스트(@)로 지정
    body.num_format_str = "@"
    header.num_format_str = "@"
    return header, body


def write_xls(path: Path, rows: List[List[str]],
              meta: List[Tuple[str, str]]) -> List[str]:
    """BIFF8 .xls 작성. 65,536행을 넘으면 시트를 나눈다."""
    book = xlwt.Workbook(encoding="utf-8")
    header_style, body_style = _xlwt_styles()

    chunks = _chunk(rows, MAX_DATA_ROWS_PER_SHEET)
    names = sheet_names_for(len(chunks))

    for name, chunk in zip(names, chunks):
        sheet = book.add_sheet(name, cell_overwrite_ok=True)
        for col, title in enumerate(HEADERS):
            sheet.write(0, col, title, header_style)
            sheet.col(col).width = 256 * COLUMN_WIDTHS[col]
        for row_index, row in enumerate(chunk, start=1):
            for col, value in enumerate(row):
                sheet.write(row_index, col, cell_text(value), body_style)
        sheet.set_panes_frozen(True)          # 1행 틀고정
        sheet.set_horz_split_pos(1)
        sheet.set_remove_splits(True)

    meta_sheet = book.add_sheet(META_SHEET_NAME, cell_overwrite_ok=True)
    meta_sheet.write(0, 0, "KEY", header_style)
    meta_sheet.write(0, 1, "VALUE", header_style)
    meta_sheet.col(0).width = 256 * 20
    meta_sheet.col(1).width = 256 * 110
    all_meta = list(meta)
    if len(names) > 1:
        all_meta.append(("SHEET_COUNT", str(len(names))))
        all_meta.append(("SHEET_NAMES", ",".join(names)))
    for row_index, (key, value) in enumerate(all_meta, start=1):
        meta_sheet.write(row_index, 0, cell_text(key), body_style)
        meta_sheet.write(row_index, 1, cell_text(value), body_style)

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(str(path))
    return names


# ==========================================================================
#  .xlsx (openpyxl) — 부가
# ==========================================================================
def write_xlsx(path: Path, rows: List[List[str]],
               meta: List[Tuple[str, str]]) -> None:
    book = XlsxWorkbook()
    sheet = book.active
    sheet.title = SHEET_NAME

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    for col, title in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = "@"
        sheet.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS[col - 1]

    for row_index, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=col)
            cell.number_format = "@"          # 텍스트 서식 (숫자 자동변환 방지)
            cell.value = cell_text(value)
            cell.alignment = Alignment(vertical="top")
    sheet.freeze_panes = "A2"

    meta_sheet = book.create_sheet(META_SHEET_NAME)
    for col, title in enumerate(("KEY", "VALUE"), start=1):
        cell = meta_sheet.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.number_format = "@"
    meta_sheet.column_dimensions["A"].width = 20
    meta_sheet.column_dimensions["B"].width = 110
    for row_index, (key, value) in enumerate(meta, start=2):
        for col, value_text in enumerate((key, value), start=1):
            cell = meta_sheet.cell(row=row_index, column=col)
            cell.number_format = "@"
            cell.value = cell_text(value_text)

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(str(path))


# ==========================================================================
#  .csv (UTF-8 BOM) — 부가
# ==========================================================================
def write_csv(path: Path, rows: List[List[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.writer(fp, quoting=csv.QUOTE_ALL)
        writer.writerow(HEADERS)
        for row in rows:
            writer.writerow([cell_text(v) for v in row])


# ==========================================================================
#  진입점
# ==========================================================================
def write_excel(findings: Sequence[Finding],
                scope,
                run_id: str,
                out_dir: str | Path = "output",
                new_cnt: int = 0,
                updated_cnt: int = 0,
                truncated_cnt: int = 0,
                run_date: Optional[str] = None,
                make_xlsx: Optional[bool] = None,
                make_csv: Optional[bool] = None) -> ExcelResult:
    """CVE_YYYYMMDD.xls (+ .xlsx/.csv) 를 만든다.

    findings 는 이미 select_rows() 로 잘라 넘긴 목록이어야 한다
    (truncated_cnt 는 그때 나온 값을 그대로 받는다).
    """
    out_dir = Path(out_dir)
    rows = build_rows(findings)
    meta = build_meta(scope, run_id, total_cnt=len(rows), new_cnt=new_cnt,
                      updated_cnt=updated_cnt, truncated_cnt=truncated_cnt)

    xls_path = out_dir / build_file_name(run_date, ".xls")
    sheet_names = write_xls(xls_path, rows, meta)
    log.info("엑셀(정본) 생성: %s (%d행, 시트 %s)",
             xls_path, len(rows), ", ".join(sheet_names))

    result = ExcelResult(xls_path=xls_path, row_count=len(rows),
                         truncated_cnt=truncated_cnt, sheet_names=sheet_names)

    if make_xlsx if make_xlsx is not None else scope.output.get("make_xlsx", True):
        try:
            result.xlsx_path = out_dir / build_file_name(run_date, ".xlsx")
            write_xlsx(result.xlsx_path, rows, meta)
            log.info("엑셀(부가) 생성: %s", result.xlsx_path)
        except Exception as exc:            # 부가 산출물 실패가 정본을 막지 않는다
            log.warning(".xlsx 생성 실패(무시하고 계속): %s", exc)
            result.xlsx_path = None

    if make_csv if make_csv is not None else scope.output.get("make_csv", True):
        try:
            result.csv_path = out_dir / build_file_name(run_date, ".csv")
            write_csv(result.csv_path, rows)
            log.info("CSV(부가) 생성: %s", result.csv_path)
        except Exception as exc:
            log.warning(".csv 생성 실패(무시하고 계속): %s", exc)
            result.csv_path = None

    return result


CLEANUP_PREFIXES = (FILE_PREFIX, "mail_preview_")


def cleanup_old_files(out_dir: str | Path, retention_days: int = 90,
                      today: Optional[datetime] = None) -> List[str]:
    """보관기간이 지난 산출물을 지운다. 지운 파일명 목록을 돌려준다.

    대상: CVE_YYYYMMDD.(xls|xlsx|csv) 와 mail_preview_YYYYMMDDHHmmss.(html|txt)
    """
    out_dir = Path(out_dir)
    if not out_dir.exists() or retention_days <= 0:
        return []
    today = today or datetime.now()
    removed: List[str] = []
    for path in out_dir.iterdir():
        if not path.is_file():
            continue
        prefix = next((p for p in CLEANUP_PREFIXES if path.name.startswith(p)), None)
        if prefix is None:
            continue
        stem = path.stem[len(prefix):][:8]
        if not stem.isdigit() or len(stem) != 8:
            continue
        try:
            file_date = datetime.strptime(stem, "%Y%m%d")
        except ValueError:
            continue
        if (today - file_date).days > retention_days:
            path.unlink()
            removed.append(path.name)
    if removed:
        log.info("보관기간(%d일) 초과 파일 %d개 삭제: %s",
                 retention_days, len(removed), ", ".join(removed[:5]))
    return removed
